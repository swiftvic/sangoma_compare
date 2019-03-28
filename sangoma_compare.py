# sangoma_compare.py
# Highlight all assemblies we are testing
import openpyxl
import re                                                                                     # Regular Expressions library
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Settings and variables
#wb1_filepath = 'Mara Transfer List_MCO Update.xlsx'
wb2_filepath = 'C:/Users/vauyeung/Documents/GitHub/Sangoma_compare/Master Test Matrix_CCD_2019.xlsx'
wb1_filepath = 'Untitled.xlsx'

#wb1_ws = {"Transfer list" : 2}
#wb2_ws = {"ProdData_TW_SW_MTP": 1, "Test Categories Mapping": 1}
wb1_ws = {"Sheet1" : 3}
wb2_ws = {"ProdData_TW_SW_MTP": 1, "Test Categories Mapping": 1}

# Flag to highlight all matches even after finding it the first instance
find_all_match = True

def open_files(wb1, wb2):
    '''
    Opens workbook1 and workbook2 file paths and assigns to workbook1 and workbook2
    '''
    wb1_path = wb1
    wb2_path = wb2

    workbook1 = openpyxl.load_workbook(wb1_path)
    workbook2 = openpyxl.load_workbook(wb2_path)

    #workbook1_ws = workbook1[ws1]                  # Opens the worksheet of workbook1
    #workbook2_ws = workbook2[ws2]                  # Opens the worksheet of workbook2

    return workbook1, workbook2

def test(org_sheet, new_sheet):
    print(org_sheet['A3'].value, new_sheet['B5'].value)

def stats(ws):
    '''
    Pass in the worksheet and will print stats of the max column and row of sheet.
    '''
    max_row = ws.max_row
    max_col = ws.max_column

    print("There are " + str(max_row) + " rows and " + str(max_col) + " columns in " + str(ws) + ".")

def color_to_hex(color_name):

    color_hex = {
        'purple':'CA19EC', 
        'red':'F2252D', 
        'yellow':'F2ED25', 
        'cyan':'25ECF2', 
        'blue':'1110FA', 
        'orange':'FAA410'
        }
    try:
        return color_hex[color_name.lower()]
    except:
        print("Invalid color: " + str(color_name))

def compare(ws1, ws1_column, ws2, ws2_column, find_all_match, ws1row_s=1, ws1row_e='max', ws2row_s=1, ws2row_e='max', color1='yellow', color2='purple'):
    '''
    ws1 = worksheet 1, ws1_column = worksheet 1 column number, ws2 = worksheet 2, ws2_column = worksheet 2 column number
    find_all_match = finds all matches and does not stop at first match
    ws1row_s = worksheet 1 start row; default = 1
    ws1row_e = worksheet 1 end row; default = max
    ws2row_s = worksheet 2 start row; default = 1
    ws2row_e = worksheet 2 end row; default = max     
    color1 = highlight color of worksheet 1 default yellow
    color2 = highlight color of worksheet 2 default purple
    Takes in worksheet 1, worksheet 1 column and compares against workseet 2, with worksheet 2 column.
    Worksheet 1 will be highlighted with matches NOT found in color1
    Worksheet 2 will be highlighted with the matches in color2.
    '''
    print("Comparing " + str(ws1) + " against " + str(ws2) + ":")
    
    # Sets values for what row to start and end for both worksheets
    # + 1 is added to include the number the user enters since range(x, ws1row_e)
    if int(ws1row_s) == 1:
        ws1row_start = 1
    else:
        ws1row_start = int(ws1row_s)
    
    if ws1row_e.lower() == 'max':
        ws1row_end = ws1.max_row + 1
    else:
        ws1row_end = int(ws1row_e) + 1          # need to +1 to include since using range(x, ws1row_e)
    
    if ws2row_s == 1:
        ws2row_start = 1
    else:
        ws2row_start = int(ws2row_s)
    
    if ws2row_e.lower() == 'max':
        ws2row_end = ws2.max_row + 1
    else:
        ws2row_end = int(ws2row_e) + 1          # need to +1 to include since using range(x, ws2row_e)

    # ws1 or worksheet1 max rows and columns
    max_ws1_row = ws1.max_row
    max_ws1_col = ws1.max_column

    # ws2 or worksheet2 max rows and columns
    max_ws2_row = ws2.max_row
    max_ws2_col = ws2.max_column

    not_found = False                                                # Flag for finding value
    highlight_color1 = color_to_hex(color1) 
    highlight_color2 = color_to_hex(color2)

    for r in range(1, max_ws1_row + 1):                              # Loop through rows, start at row 1, and max row + 1 to include last row
        ws1_value = ws1.cell(r, ws1_column).value
        find_count = 0

        for ws2_r in range(1, max_ws2_row + 1):                     # max row + 1 to include the last line
            if ws1_value == ws2.cell(ws2_r, ws2_column).value: 
                ws2.cell(ws2_r, ws2_column).fill = PatternFill(patternType='solid', fill_type='solid', fgColor=highlight_color2) # Highlight cell
                not_found = False
                find_count += 1
                #print(ws1_value)                                # Prints out found value for debugging
                if not find_all_match:                           # If user wants to find all matches even after finding the first instance
                    break                                        # Found value in wb2 sheet, break out, don't break out to highlight duplicates as well
            elif ws1_value == None:
                not_found = False
                break
            else:
                pass
                not_found = True
                      
        if not_found and find_count == 0:
            print(str(ws1_value) + " not found.")
            ws1.cell(r, ws1_column).fill = PatternFill(patternType='solid', fill_type='solid', fgColor=highlight_color1)   # Highlight cell
            find_count = 0
        elif find_count > 1:
            print(str(ws1_value) + " found " + str(find_count) + " times.")             # Displays duplicates found.
        else:
            pass

def parse_filename(wb_filepath):
    '''
    Parses filepath from '/' and '.xlsx' of workbook and leave only file name.
    Returns str of filename only.
    '''
    wb_parse_slash = wb_filepath.split('/')[-1]
    wb_filename_only = wb_parse_slash.split('.')[0]

    return str(wb_filename_only)

if __name__ == '__main__':
    wb1, wb2 = open_files(wb1_filepath, wb2_filepath)

    for wb1_ws_key in wb1_ws:                                                                       # Loop through each worksheet in workbook 1 (usually just 1)
        #print(str(wb1_ws_key))
        for wb2_ws_key in wb2_ws:                                                                   # Loop through each worksheet in workbook 2
            '''
            Feeds in workseet 1 key (worksheet title) and worksheet 1 value (worksheet column) and
            feeds in workseet 2 key (worksheet title) and worksheet 2 value (worksheet column) into 
            compare function so it can compare worksheet 1 with multiple workseets of workbook 2.
            '''
            compare(wb1[wb1_ws_key], wb1_ws[wb1_ws_key], wb2[wb2_ws_key], wb2_ws[wb2_ws_key], find_all_match)      
    #test(ws1, ws2)
    #print(ws1['B4'].value)
    #stats(ws1)
    #stats(ws2)
    
    # Save workbooks
    #wb1.save(parse_filename(wb1_filepath) + ' highlighted.xlsx')
    wb2.save(parse_filename(wb2_filepath) + ' highlighted.xlsx')